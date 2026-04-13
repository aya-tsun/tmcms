import csv
import io
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_db
from ..models.user import User
from ..models.material import Material
from ..models.tag import MaterialTag
from ..auth.jwt import get_current_user
from .materials import _build_material_out

router = APIRouter(prefix="/api/export", tags=["export"])


def _get_filtered_materials(
    db: Session,
    search: Optional[str],
    provider: Optional[str],
    category: Optional[str],
    tag_ids: Optional[str],
    material_ids: Optional[str],
):
    from sqlalchemy import or_

    query = db.query(Material)
    if material_ids:
        ids = [int(x) for x in material_ids.split(",") if x.strip().isdigit()]
        if ids:
            query = query.filter(Material.id.in_(ids))
    else:
        if search:
            query = query.filter(
                or_(
                    Material.name.ilike(f"%{search}%"),
                    Material.provider.ilike(f"%{search}%"),
                )
            )
        if provider:
            query = query.filter(Material.provider.ilike(f"%{provider}%"))
        if category:
            query = query.filter(Material.category.ilike(f"%{category}%"))
        if tag_ids:
            ids = [int(x) for x in tag_ids.split(",") if x.strip().isdigit()]
            for tid in ids:
                query = query.filter(
                    Material.id.in_(
                        db.query(MaterialTag.material_id).filter(MaterialTag.tag_id == tid)
                    )
                )
    return query.all()


HEADERS = [
    "ID", "教材名", "提供元", "カテゴリ", "URL",
    "受講時間(時間)", "費用(円)", "対象レベル", "言語",
    "総合評価", "コンテンツの質", "わかりやすさ", "費用対効果",
    "タグ", "登録日", "登録者",
]


def _material_row(m):
    tags = ", ".join(t.name for t in [mt.tag for mt in m.material_tags])
    evals = m.evaluations
    overall = round(sum(e.overall_score for e in evals) / len(evals), 2) if evals else ""
    quality = round(sum(e.quality for e in evals if e.quality) / len([e for e in evals if e.quality]), 2) if any(e.quality for e in evals) else ""
    clarity = round(sum(e.clarity for e in evals if e.clarity) / len([e for e in evals if e.clarity]), 2) if any(e.clarity for e in evals) else ""
    cost_eff = round(sum(e.cost_effectiveness for e in evals if e.cost_effectiveness) / len([e for e in evals if e.cost_effectiveness]), 2) if any(e.cost_effectiveness for e in evals) else ""
    creator_name = m.creator.name if m.creator else ""
    return [
        m.id, m.name, m.provider, m.category, m.url,
        m.duration or "", m.cost or "", m.level or "", m.language or "",
        overall, quality, clarity, cost_eff,
        tags,
        m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
        creator_name,
    ]


@router.get("/csv")
def export_csv(
    search: Optional[str] = Query(None),
    provider: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    tag_ids: Optional[str] = Query(None),
    material_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    materials = _get_filtered_materials(db, search, provider, category, tag_ids, material_ids)
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata
    writer.writerow(["出力日時", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    filters = []
    if search:
        filters.append(f"検索: {search}")
    if provider:
        filters.append(f"提供元: {provider}")
    if category:
        filters.append(f"カテゴリ: {category}")
    writer.writerow(["フィルタ条件", " | ".join(filters) if filters else "なし"])
    writer.writerow([])

    writer.writerow(HEADERS)
    for m in materials:
        writer.writerow(_material_row(m))

    output.seek(0)
    content = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=tmcms_export_{now}.csv"},
    )


@router.get("/xlsx")
def export_xlsx(
    search: Optional[str] = Query(None),
    provider: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    tag_ids: Optional[str] = Query(None),
    material_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    materials = _get_filtered_materials(db, search, provider, category, tag_ids, material_ids)
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "教材一覧"

    # Metadata rows
    ws.append(["出力日時", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    filters = []
    if search:
        filters.append(f"検索: {search}")
    if provider:
        filters.append(f"提供元: {provider}")
    if category:
        filters.append(f"カテゴリ: {category}")
    ws.append(["フィルタ条件", " | ".join(filters) if filters else "なし"])
    ws.append([])

    # Header row
    header_row = ws.max_row + 1
    ws.append(HEADERS)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for m in materials:
        ws.append(_material_row(m))

    # Auto-fit columns approximately
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tmcms_export_{now}.xlsx"},
    )
